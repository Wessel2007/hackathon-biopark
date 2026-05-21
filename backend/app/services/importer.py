import re
from datetime import datetime, date

import openpyxl
import pandas as pd

from app.supabase_client import SupabaseClient

SKIP_SHEETS = {"Configuração"}

STATUS_MAP = {
    "CANC": "CANCELADO",
    "ANÁL": "EM ANDAMENTO",
    "APRO": "APROVADO",
}

# Separadores aceitos entre Projeto e Protocolo, em ordem de prioridade.
# Apenas variantes com espaço ou em-dash explícito — evita falsos positivos
# com hifens dentro do número do protocolo (ex: "SEI-12345").
_SEPS = [" – ", " - ", " / ", "– "]

_CAMPOS_OBRIGATORIOS = ("status", "projeto", "protocolo", "atividade")
_CAMPOS_TEXTO = ("orgao_site_consultado", "atribuido_a", "situacao", "anotacoes", "url_consulta", "observacao_consulta")
_CAMPOS_DATA = ("data_abertura", "data_finalizacao")

# Limites do schema Supabase. Após rodar migration_widen_protocol_text.sql,
# projeto/orgao passam a 500 e atividade vira TEXT (use 10000).
_LIMITES_CAMPO = {
    "status": 50,
    "projeto": 500,
    "protocolo": 100,
    "atividade": 10000,
    "orgao_site_consultado": 500,
    "atribuido_a": 100,
    "situacao": 100,
    "url_consulta": 500,
    "anotacoes": 10000,
}

# Fallback para bancos ainda com VARCHAR(200) — evita erro 22001 sem migration
_LIMITES_CAMPO_LEGACY = {
    **{k: v for k, v in _LIMITES_CAMPO.items()},
    "projeto": 200,
    "atividade": 200,
    "orgao_site_consultado": 200,
}


# ---------------------------------------------------------------------------
# Helpers de transformação
# ---------------------------------------------------------------------------

def _normaliza_status(s):
    if not s:
        return ""
    s = str(s).strip().upper()
    s = re.sub(r"[.\s]+$", "", s)
    return STATUS_MAP.get(s, s)


def _limpa_texto(v):
    if v is None:
        return None
    s = str(v).strip().replace("\n", " ").replace("\r", " ")
    s = re.sub(r" {2,}", " ", s)
    return s if s else None


def _truncar_campo(campo: str, valor: str | None, limites: dict | None = None) -> tuple[str | None, str | None]:
    """Corta texto ao limite do banco. Retorna (valor, aviso) — aviso só se houve corte."""
    if valor is None:
        return None, None
    limite = (limites or _LIMITES_CAMPO).get(campo)
    if not limite or len(valor) <= limite:
        return valor, None
    return valor[:limite], f"{campo} truncado ({len(valor)} → {limite} caracteres)"


def _aplicar_limites_payload(payload: dict, limites: dict | None = None) -> list[str]:
    """Trunca campos de texto do payload; retorna lista de avisos."""
    avisos = []
    for campo in (*_CAMPOS_OBRIGATORIOS, *_CAMPOS_TEXTO):
        if campo not in payload:
            continue
        val = payload.get(campo)
        if val is None:
            continue
        novo, aviso = _truncar_campo(campo, str(val), limites)
        if aviso:
            avisos.append(aviso)
        payload[campo] = novo if campo == "orgao_site_consultado" else (novo or None)
    return avisos


def _normaliza_projeto(nome: str) -> str:
    if not nome:
        return nome
    nome = re.sub(r" {2,}", " ", nome.strip())
    return nome.title()


def _formata_data(v):
    """Converte qualquer valor de data para ISO string ou None — nunca lança exceção."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    s = str(v).strip()
    if not s or s.lower() in ("nan", "none", "nat"):
        return None
    try:
        return pd.to_datetime(s).date().isoformat()
    except Exception:
        return None


def _inferir_ativo(status, data_finalizacao):
    if str(status or "").upper() == "CANCELADO":
        return False
    if data_finalizacao:
        return False
    return True


def _projeto_do_nome(nome_aba):
    partes = nome_aba.split(" - ", 1)
    raw = partes[-1].strip() if len(partes) > 1 else nome_aba.strip()
    return _normaliza_projeto(raw)


def _safe_get(row, idx, default=None):
    if idx is None:
        return default
    try:
        return row[idx]
    except (IndexError, TypeError):
        return default


def _extrai_projeto_protocolo(col_b_val, col_b_header, nome_aba):
    """
    Separa o campo combinado "Projeto + Protocolo".
    Aceita os separadores: ' – ', ' - ', ' / ', '– ' (em ordem).
    Se nenhum separador for encontrado, usa o nome da aba como projeto.
    """
    col_b_str = str(col_b_header or "").lower()
    if "projeto" not in col_b_str:
        return _projeto_do_nome(nome_aba), _limpa_texto(col_b_val) or ""

    texto = _limpa_texto(col_b_val) or ""
    for sep in _SEPS:
        if sep in texto:
            partes = texto.split(sep, 1)
            projeto, protocolo = partes[0].strip(), partes[1].strip()
            if projeto and protocolo:
                return _normaliza_projeto(projeto), protocolo

    # Sem separador reconhecível — projeto do nome da aba, tudo vira protocolo
    return _projeto_do_nome(nome_aba), texto


def _revalidar_payload(reg: dict):
    """
    Revalida e sanitiza um registro recebido do frontend antes de inserir.
    Garante que campos obrigatórios existem, datas são válidas e tipos estão corretos.
    Retorna (payload_limpo, msg_erro). Se msg_erro não for None, a linha deve ser ignorada.
    """
    payload = {}

    for campo in _CAMPOS_OBRIGATORIOS:
        val = str(reg.get(campo, "") or "").strip()
        if campo == "status":
            val = _normaliza_status(val)
        elif campo == "projeto":
            val = _normaliza_projeto(val)
        if not val:
            return None, f"Campo obrigatório ausente: {campo}"
        payload[campo] = val

    for campo in _CAMPOS_TEXTO:
        val = str(reg.get(campo, "") or "").strip()
        # orgao_site_consultado é NOT NULL no banco; demais campos aceitam NULL
        payload[campo] = val if campo == "orgao_site_consultado" else (val or None)

    for campo in _CAMPOS_DATA:
        payload[campo] = _formata_data(reg.get(campo))

    if not payload["data_abertura"]:
        return None, "data_abertura ausente ou inválida"

    ativo = reg.get("ativo")
    if isinstance(ativo, bool):
        payload["ativo"] = ativo
    elif isinstance(ativo, str):
        payload["ativo"] = ativo.lower() in ("true", "sim", "yes", "1")
    else:
        payload["ativo"] = bool(ativo) if ativo is not None else True

    anotacoes = _limpa_texto(reg.get("anotacoes"))
    if anotacoes:
        payload["anotacoes"] = anotacoes

    if not (payload.get("orgao_site_consultado") or "").strip():
        return None, "Campo obrigatório ausente: orgao_site_consultado"
    if payload["ativo"] and not (payload.get("atribuido_a") or "").strip():
        return None, "Campo obrigatório ausente: atribuido_a (protocolo ativo)"

    avisos = _aplicar_limites_payload(payload, _LIMITES_CAMPO)
    if avisos:
        return payload, "; ".join(avisos)
    return payload, None


# ---------------------------------------------------------------------------
# Detecção de colunas (arquivo-base)
# ---------------------------------------------------------------------------

def _detectar_indices(cabecalho):
    idx = {}
    for i, v in enumerate(cabecalho):
        if v is None:
            continue
        nome = str(v).lower().replace("\n", " ").strip()
        if "status" in nome and i == 0:
            idx["status"] = i
        elif ("projeto" in nome and "protocolo" in nome) or (nome == "protocolo" and i == 1):
            idx["col_b"] = i
            idx["col_b_header"] = v
        elif "atividade" in nome:
            idx["atividade"] = i
        elif "atribu" in nome:
            idx["atribuido_a"] = i
        elif "real" in nome and ("início" in nome or "inicio" in nome or "nicio" in nome):
            idx["data_abertura"] = i
        elif "real" in nome and ("término" in nome or "termino" in nome or "rmino" in nome):
            idx["data_finalizacao"] = i
        elif "anota" in nome:
            idx["anotacoes"] = i
        elif "situa" in nome:
            idx["situacao"] = i
        elif "aprova" in nome:
            idx["orgao"] = i
    return idx


# ---------------------------------------------------------------------------
# Parsers por formato
# ---------------------------------------------------------------------------

def _processar_aba_base(ws, nome_aba):
    """Parse uma aba do arquivo-base (cabeçalho na linha 4, dados a partir da linha 5)."""
    registros, ignorados, erros = [], [], []

    try:
        cab = list(next(ws.iter_rows(min_row=4, max_row=4, values_only=True)))
    except StopIteration:
        erros.append({"linha": f"{nome_aba}:4", "erro": "Aba sem cabeçalho na linha 4"})
        return registros, ignorados, erros

    idx = _detectar_indices(cab)
    col_b_header = idx.get("col_b_header", cab[1] if len(cab) > 1 else "")

    for row_num, row in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
        if not any(c for c in row if c is not None):
            continue

        linha = f"{nome_aba}:{row_num}"
        try:
            status_raw = _safe_get(row, idx.get("status", 0))
            if not status_raw:
                ignorados.append({"linha": linha, "motivo": "Status ausente"})
                continue

            status = _normaliza_status(status_raw)
            col_b_val = _safe_get(row, idx.get("col_b", 1))
            projeto, protocolo = _extrai_projeto_protocolo(col_b_val, col_b_header, nome_aba)

            atividade = _limpa_texto(_safe_get(row, idx.get("atividade", 2)))
            orgao = _limpa_texto(_safe_get(row, idx.get("orgao", 3)))
            atribuido_a = _limpa_texto(_safe_get(row, idx.get("atribuido_a", 4)))
            data_abertura = _formata_data(_safe_get(row, idx.get("data_abertura", 8)))
            data_finalizacao = _formata_data(_safe_get(row, idx.get("data_finalizacao", 9)))
            situacao = _limpa_texto(_safe_get(row, idx.get("situacao")))
            anotacoes = _limpa_texto(_safe_get(row, idx.get("anotacoes")))
            ativo = _inferir_ativo(status, data_finalizacao)

            if not all([status, projeto, protocolo, atividade]):
                ignorados.append({"linha": linha, "motivo": "Campos obrigatórios ausentes"})
                continue

            if not data_abertura:
                ignorados.append({"linha": linha, "motivo": "data_abertura ausente ou inválida"})
                continue

            if not orgao:
                ignorados.append({"linha": linha, "motivo": "Órgão / site consultado ausente"})
                continue

            if ativo and not atribuido_a:
                ignorados.append({"linha": linha, "motivo": "Atribuído a ausente (protocolo ativo)"})
                continue

            registros.append({
                "linha": linha,
                "status": status,
                "projeto": projeto,
                "protocolo": protocolo,
                "atividade": atividade,
                "orgao_site_consultado": orgao,
                "atribuido_a": atribuido_a,
                "data_abertura": data_abertura,
                "data_finalizacao": data_finalizacao,
                "situacao": situacao,
                "anotacoes": anotacoes,
                "ativo": ativo,
                "url_consulta": None,
                "observacao_consulta": None,
            })
        except Exception as e:
            erros.append({"linha": linha, "erro": str(e)})

    return registros, ignorados, erros


def _cel(row, col, default=""):
    """Lê célula do pandas evitando que NaN vire a string 'nan'."""
    v = row.get(col, default)
    try:
        if pd.isna(v):
            return default
    except (TypeError, ValueError):
        pass
    return v


def _processar_carga_inicial(file_buffer):
    """Parse a aba 'Carga Inicial' (formato de saída com colunas já separadas)."""
    registros, ignorados, erros = [], [], []

    try:
        df = pd.read_excel(file_buffer, sheet_name="Carga Inicial", header=3)
    except Exception as e:
        return None, [], [], f"Erro ao ler aba 'Carga Inicial': {str(e)}"

    df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]

    for i, row in df.iterrows():
        linha = f"Carga Inicial:{i + 5}"
        try:
            status = _normaliza_status(_cel(row, "status"))
            projeto = _normaliza_projeto(str(_cel(row, "projeto")).strip())
            protocolo = str(_cel(row, "protocolo")).strip()
            atividade = str(_cel(row, "atividade")).strip()
            orgao = str(_cel(row, "orgao_site_consultado")).strip()

            if not all([status, projeto, protocolo, atividade, orgao]):
                ignorados.append({"linha": linha, "motivo": "Campos obrigatórios ausentes"})
                continue

            data_abertura = _formata_data(row.get("data_abertura"))
            if not data_abertura:
                ignorados.append({"linha": linha, "motivo": "data_abertura ausente ou inválida"})
                continue

            data_finalizacao = _formata_data(row.get("data_finalizacao"))
            ativo_raw = str(_cel(row, "ativo", "Sim")).strip().lower()
            ativo = ativo_raw in ("sim", "yes", "true", "1")
            atribuido_a = str(_cel(row, "atribuido_a")).strip() or None
            anotacoes = str(_cel(row, "anotacoes")).strip() or None

            if ativo and not atribuido_a:
                ignorados.append({"linha": linha, "motivo": "Atribuído a ausente (protocolo ativo)"})
                continue

            registros.append({
                "linha": linha,
                "status": status,
                "projeto": projeto,
                "protocolo": protocolo,
                "atividade": atividade,
                "orgao_site_consultado": orgao,
                "atribuido_a": atribuido_a,
                "data_abertura": data_abertura,
                "data_finalizacao": data_finalizacao,
                "situacao": str(_cel(row, "situacao")).strip() or None,
                "anotacoes": anotacoes,
                "ativo": ativo,
                "url_consulta": str(_cel(row, "url_consulta")).strip() or None,
                "observacao_consulta": str(_cel(row, "observacao_consulta")).strip() or None,
            })
        except Exception as e:
            erros.append({"linha": linha, "erro": str(e)})

    return registros, ignorados, erros, None


# ---------------------------------------------------------------------------
# API pública
# ---------------------------------------------------------------------------

def parse_spreadsheet(file_buffer) -> dict:
    """
    Parse o arquivo .xlsx e retorna {rows, ignorados, erros} sem tocar no banco.
    Detecta formato automaticamente:
    - Aba 'Carga Inicial' presente → formato de saída (colunas separadas)
    - Sem essa aba → arquivo-base multi-abas ('Projeto – Protocolo' combinados)
    """
    try:
        wb = openpyxl.load_workbook(file_buffer, data_only=True)
        sheet_names = wb.sheetnames
        wb.close()
    except Exception as e:
        return {"error": f"Erro ao abrir arquivo: {str(e)}"}

    file_buffer.seek(0)

    if "Carga Inicial" in sheet_names:
        registros, ignorados, erros, error = _processar_carga_inicial(file_buffer)
        if error:
            return {"error": error}
    else:
        try:
            wb = openpyxl.load_workbook(file_buffer, data_only=True)
        except Exception as e:
            return {"error": f"Erro ao abrir arquivo: {str(e)}"}

        registros, ignorados, erros = [], [], []
        for nome in wb.sheetnames:
            if nome in SKIP_SHEETS:
                continue
            try:
                r, ig, er = _processar_aba_base(wb[nome], nome)
                registros.extend(r)
                ignorados.extend(ig)
                erros.extend(er)
            except Exception as e:
                erros.append({"linha": f"{nome}:?", "erro": f"Erro ao processar aba: {str(e)}"})
        wb.close()

    return {"rows": registros, "ignorados": ignorados, "erros": erros}


def _chave_protocolo(payload: dict) -> tuple[str, str]:
    """Chave única no banco: constraint uq_projeto_protocolo."""
    return (payload["projeto"], payload["protocolo"])


def _inserir_rows(rows: list, sb: SupabaseClient):
    """
    Revalida, checa duplicidade (projeto + protocolo) e insere em lote.
    Usa 2 requests ao banco independente do número de linhas:
    1 SELECT para buscar existentes, 1 INSERT em lote com os novos.
    """
    imported, ignorados, erros = [], [], []

    # Passo 1: validar todos os registros em memória
    payloads_validos: list[tuple[str, dict]] = []
    for reg in rows:
        linha = reg.get("linha", "?")
        payload, erro_validacao = _revalidar_payload(reg)
        if payload is None:
            ignorados.append({"linha": linha, "motivo": erro_validacao})
        else:
            if erro_validacao:
                ignorados.append({"linha": linha, "motivo": erro_validacao})
            payloads_validos.append((linha, payload))

    if not payloads_validos:
        return imported, ignorados, erros

    # Passo 2: buscar chaves (projeto, protocolo) já cadastradas
    try:
        existing_raw = (
            sb.table("protocols")
            .select("projeto,protocolo")
            .execute()
        )
        existing = {
            (r["projeto"], r["protocolo"])
            for r in (existing_raw.data or [])
        }
    except Exception as e:
        erros.append({"linha": "dedup-check", "erro": f"Erro ao verificar duplicatas: {str(e)}"})
        return imported, ignorados, erros

    # Passo 3: filtrar duplicatas (banco + repetidas na mesma planilha)
    to_insert: list[tuple[str, dict]] = []
    for linha, payload in payloads_validos:
        key = _chave_protocolo(payload)
        if key in existing:
            ignorados.append({
                "linha": linha,
                "motivo": f"Duplicata ignorada: {payload['projeto']} / {payload['protocolo']} já cadastrado",
            })
        else:
            to_insert.append((linha, payload))
            existing.add(key)

    if not to_insert:
        return imported, ignorados, erros

    # Passo 4: insert em lote (com fallback se o banco ainda usa VARCHAR(200))
    payloads_insert = [p for _, p in to_insert]
    try:
        sb.table("protocols").insert(payloads_insert).execute()
        imported = [
            {"linha": linha, "protocolo": p["protocolo"], "projeto": p["projeto"]}
            for linha, p in to_insert
        ]
    except Exception as e:
        err_msg = str(e)
        if "22001" in err_msg or "value too long" in err_msg.lower():
            for p in payloads_insert:
                _aplicar_limites_payload(p, _LIMITES_CAMPO_LEGACY)
            try:
                sb.table("protocols").insert(payloads_insert).execute()
                imported = [
                    {"linha": linha, "protocolo": p["protocolo"], "projeto": p["projeto"]}
                    for linha, p in to_insert
                ]
                erros.append({
                    "linha": "bulk-insert",
                    "erro": (
                        "Importado com textos truncados (colunas curtas no banco). "
                        "Execute backend/migration_widen_protocol_text.sql no Supabase para preservar textos longos."
                    ),
                })
            except Exception as e2:
                erros.append({"linha": "bulk-insert", "erro": f"Erro no insert em lote: {str(e2)}"})
        elif "23505" in err_msg or "duplicate key" in err_msg.lower() or "409" in err_msg:
            for linha, p in to_insert:
                try:
                    sb.table("protocols").insert(p).execute()
                    imported.append({
                        "linha": linha,
                        "protocolo": p["protocolo"],
                        "projeto": p["projeto"],
                    })
                    existing.add(_chave_protocolo(p))
                except Exception as row_err:
                    row_msg = str(row_err)
                    if "23505" in row_msg or "duplicate key" in row_msg.lower() or "409" in row_msg:
                        ignorados.append({
                            "linha": linha,
                            "motivo": f"Duplicata ignorada: {p['projeto']} / {p['protocolo']} já cadastrado",
                        })
                    else:
                        erros.append({"linha": linha, "erro": row_msg})
        else:
            erros.append({"linha": "bulk-insert", "erro": f"Erro no insert em lote: {err_msg}"})

    return imported, ignorados, erros


def import_spreadsheet(file_buffer, sb: SupabaseClient) -> dict:
    """Parse + insere em uma etapa só (endpoint legado /spreadsheet)."""
    parsed = parse_spreadsheet(file_buffer)
    if "error" in parsed:
        return parsed

    imported, ignorados_new, erros_new = _inserir_rows(parsed["rows"], sb)
    return {
        "importados": imported,
        "ignorados": parsed["ignorados"] + ignorados_new,
        "erros": parsed["erros"] + erros_new,
    }


def confirm_import(rows: list, sb: SupabaseClient) -> dict:
    """Revalida e insere as linhas enviadas pelo frontend após o preview."""
    imported, ignorados, erros = _inserir_rows(rows, sb)
    return {"importados": imported, "ignorados": ignorados, "erros": erros}
