import openpyxl
from openpyxl import load_workbook
from datetime import datetime, date
import re
import warnings
warnings.filterwarnings("ignore")

BASE_PATH   = "Planilhas/Desafio 4 - Base para carga inicial de protocolos.xlsx"
MODELO_PATH = "Planilhas/Desafio 4 - Modelo Carga Inicial Protocolos.xlsx"
OUTPUT_PATH = "Planilhas/Desafio 4 - Carga Inicial Gerada.xlsx"

# Abas que não são de protocolos
SKIP_SHEETS = {"Configuração"}

STATUS_MAP = {
    "CANC": "CANCELADO",
    "ANÁL": "EM ANDAMENTO",
    "ANÁL": "EM ANDAMENTO",
}

def normaliza_status(s):
    if not s:
        return ""
    s = str(s).strip().upper()
    return STATUS_MAP.get(s, s)

def limpa_texto(v):
    if v is None:
        return None
    s = str(v).strip().replace("\n", " ").replace("\r", " ")
    s = re.sub(r" {2,}", " ", s)
    return s if s else None

def formata_data(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    return s if s else None

def extrai_projeto_protocolo(valor_col_b, header_col_b, nome_aba):
    """
    Se o cabeçalho da col B contém 'Projeto', faz split por ' - '.
    Caso contrário (coluna só de Protocolo), projeto vem do nome da aba.
    """
    col_b_str = str(header_col_b or "").lower()
    usa_campo_combinado = "projeto" in col_b_str

    if usa_campo_combinado:
        raw = limpa_texto(valor_col_b) or ""
        if " - " in raw:
            partes = raw.split(" - ", 1)
            projeto   = partes[0].strip()
            protocolo = partes[1].strip()
        else:
            # Sem separador — coloca tudo em protocolo, projeto do nome da aba
            projeto   = _projeto_do_nome(nome_aba)
            protocolo = raw
    else:
        projeto   = _projeto_do_nome(nome_aba)
        protocolo = limpa_texto(valor_col_b) or ""

    return projeto, protocolo

def _projeto_do_nome(nome_aba):
    """Extrai o nome do empreendimento do nome da aba: 'Prot. TLD - Horizont' → 'Horizont'"""
    partes = nome_aba.split(" - ", 1)
    return partes[-1].strip() if len(partes) > 1 else nome_aba.strip()

def inferir_ativo(status, data_finalizacao):
    s = str(status or "").upper()
    if s == "CANCELADO":
        return "Não"
    if data_finalizacao:
        return "Não"
    return "Sim"

def detectar_indices(cabecalho):
    """
    Retorna índices das colunas relevantes baseado no cabeçalho real.
    Lida com abas de 14 ou 15 colunas.
    """
    idx = {}
    for i, v in enumerate(cabecalho):
        if v is None:
            continue
        nome = str(v).lower().replace("\n", " ").strip()
        if "status" in nome and i == 0:
            idx["status"] = i
        elif ("projeto" in nome and "protocolo" in nome) or (nome == "protocolo" and i == 1):
            idx["col_b"] = i
            idx["col_b_header"] = v
        elif "atividade" in nome:
            idx["atividade"] = i
        elif "atribu" in nome:
            idx["atribuido_a"] = i
        elif "real" in nome and ("início" in nome or "inicio" in nome or nome.endswith("início") or "nicio" in nome):
            idx["data_abertura"] = i
        elif "real" in nome and ("término" in nome or "termino" in nome or "rmino" in nome):
            idx["data_finalizacao"] = i
        elif "anota" in nome:
            idx["anotacoes"] = i
        elif "situa" in nome:
            idx["situacao"] = i
        elif "aprova" in nome:
            idx["orgao"] = i
    return idx

def processar_aba(ws, nome_aba):
    registros = []

    # Cabeçalho sempre na linha 4
    cab = [c.value for c in ws[4]]
    idx = detectar_indices(cab)

    col_b_header = idx.get("col_b_header", cab[1] if len(cab) > 1 else "")

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        # Pula linhas totalmente vazias
        if not any(c for c in row if c is not None):
            continue

        status_raw = row[idx.get("status", 0)]
        if not status_raw:
            continue

        status = normaliza_status(status_raw)

        col_b_val = row[idx.get("col_b", 1)]
        projeto, protocolo = extrai_projeto_protocolo(col_b_val, col_b_header, nome_aba)

        atividade    = limpa_texto(row[idx.get("atividade", 2)])
        orgao        = limpa_texto(row[idx.get("orgao", 3)])
        atribuido_a  = limpa_texto(row[idx.get("atribuido_a", 4)])
        data_aber    = formata_data(row[idx.get("data_abertura", 8)])
        data_fim     = formata_data(row[idx.get("data_finalizacao", 9)])
        anotacoes    = limpa_texto(row[idx.get("anotacoes")] if "anotacoes" in idx else None)
        situacao     = limpa_texto(row[idx.get("situacao")] if "situacao" in idx else None)

        ativo = inferir_ativo(status, data_fim)

        registros.append({
            "status":               status,
            "projeto":              projeto,
            "protocolo":            protocolo,
            "atividade":            atividade,
            "orgao_site_consultado": orgao,
            "atribuido_a":          atribuido_a,
            "data_abertura":        data_aber,
            "data_finalizacao":     data_fim,
            "situacao":             situacao,
            "anotacoes":            anotacoes,
            "ativo":                ativo,
            "url_consulta":         None,
            "ultima_consulta":      None,
            "observacao_consulta":  None,
        })

    return registros

def gerar_carga():
    wb_base   = load_workbook(BASE_PATH, data_only=False)
    wb_modelo = load_workbook(MODELO_PATH)

    ws_saida = wb_modelo["Carga Inicial"]

    # Cabeçalho do modelo está na linha 4, dados começam na linha 5
    COLUNAS = [
        "status", "projeto", "protocolo", "atividade",
        "orgao_site_consultado", "atribuido_a",
        "data_abertura", "data_finalizacao",
        "situacao", "anotacoes", "ativo",
        "url_consulta", "ultima_consulta", "observacao_consulta",
        # duracao_calculada_dias ignorada
    ]

    todos = []
    for nome in wb_base.sheetnames:
        if nome in SKIP_SHEETS:
            continue
        ws = wb_base[nome]
        registros = processar_aba(ws, nome)
        for r in registros:
            r["_aba"] = nome
        todos.extend(registros)
        print(f"  {nome}: {len(registros)} registros")

    print(f"\nTotal consolidado: {len(todos)} registros")

    # Remove linhas existentes de dados no modelo (mantém as 4 primeiras — títulos e cabeçalho)
    if ws_saida.max_row >= 5:
        ws_saida.delete_rows(5, ws_saida.max_row - 4)

    for i, reg in enumerate(todos):
        linha = 5 + i
        for j, campo in enumerate(COLUNAS):
            ws_saida.cell(row=linha, column=j + 1, value=reg.get(campo))

    wb_modelo.save(OUTPUT_PATH)
    print(f"\nArquivo salvo em: {OUTPUT_PATH}")

if __name__ == "__main__":
    gerar_carga()
